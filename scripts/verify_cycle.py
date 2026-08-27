"""Verification pass - run at the end of every cycle before the deliverables are shared.

Checks (CLAUDE.md verification rule):
  1. Row counts reconcile: raw source -> deduped -> anchors -> workbook sheets, and every
     audit row dropped along the way is attributable to a named cause.
  2. No duplicate rows in any output sheet, plus proof that Graph's chunked-event
     fragments were reassembled - they differ in the microseconds of Date, so an
     exact-match test alone never sees them.
  3. Exception totals and severity counts agree between correlation-stats and the CSVs.
  4. Every input file appears in export-manifest.json.
  5. No orphan references: every activation_id used in correlated/exception rows exists.
  6. Every correlated action falls inside its activation's stated window.
  7. Every Summary formula, recomputed from the source CSVs, agrees with correlation-stats.

Exits non-zero if any check fails, so it can gate a scheduled run.

On check 7 and cached values: the Summary sheet holds live formulas by design, and
openpyxl writes formulas without computing them - so a freshly built workbook has no
cached formula values at all. Checks that compared against those cached values used to
skip silently while still reporting PASS, which meant a tampered or stale Exceptions
sheet could clear the gate. correlation-stats-<label>.json is the authority for every
figure the reports use, so formulas are now recomputed from the CSVs and compared against
stats instead. Cached values are still checked when present (after Excel or LibreOffice
has opened the file), but nothing depends on them, and a run where no formula could be
compared to anything is a failure rather than a pass.
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd

sys.path.insert(0, str(Path(__file__).resolve().parent))

from common import month_paths, read_manifest

RESULTS: list[tuple[bool, str, str]] = []

COUNTA_RE = re.compile(r"^=COUNTA\('?([^'!]+)'?!\$?([A-Z]+)\$?\d+:\$?([A-Z]+)\$?(\d+)\)$")
COUNTIF_RE = re.compile(
    r'^=COUNTIF\(\'?([^\'!]+)\'?!\$?([A-Z]+)\$?\d+:\$?([A-Z]+)\$?(\d+),"(.*)"\)$')


def _col_index(letter: str) -> int:
    n = 0
    for ch in letter:
        n = n * 26 + (ord(ch) - 64)
    return n - 1


def eval_count_formula(formula: str, frames: dict[str, pd.DataFrame]) -> int | None:
    """Recompute a Summary COUNTA/COUNTIF straight from the source CSVs. Returns None for
    anything this does not recognise, so unknown formulas are skipped rather than guessed at."""
    m = COUNTA_RE.match(formula)
    if m:
        sheet, col, _, _ = m.groups()
        df = frames.get(sheet)
        if df is None or df.empty:
            return 0
        i = _col_index(col)
        if i >= len(df.columns):
            return None
        return int((df.iloc[:, i].astype(str).str.strip() != "").sum())

    m = COUNTIF_RE.match(formula)
    if m:
        sheet, col, _, _, criterion = m.groups()
        df = frames.get(sheet)
        if df is None or df.empty:
            return 0
        i = _col_index(col)
        if i >= len(df.columns):
            return None
        return int((df.iloc[:, i].astype(str).str.strip() == criterion).sum())
    return None


def summary_expectations(stats: dict) -> dict[str, int]:
    """Map each Summary row label to the authoritative figure in correlation-stats.

    correlation-stats-<label>.json is written by correlate.py and is the single source of
    truth for every count the reports use, so it can verify the workbook's formulas
    without the workbook needing cached formula values.

    Severity and exception-class rows are expanded from their stats dictionaries; a label
    absent from stats maps to 0, because a COUNTIF for a class with no rows should return
    0 rather than be skipped.
    """
    exp: dict[str, int] = {}

    if "activations_successful" in stats:
        exp["Successful activations"] = int(stats["activations_successful"])
    by_role = stats.get("by_role") or {}
    if by_role:
        exp["Global Administrator activations"] = int(by_role.get("Global Administrator", 0))
    if "exception_rows" in stats:
        exp["Total exceptions"] = int(stats["exception_rows"])

    by_sev = stats.get("exceptions_by_severity")
    if isinstance(by_sev, dict):
        for name in ("High", "Medium", "Low", "Informational"):
            exp[name] = int(by_sev.get(name, 0))

    by_class = stats.get("exceptions_by_class")
    if isinstance(by_class, dict):
        for name, count in by_class.items():
            exp[str(name)] = int(count)

    return exp


CLASS_LABEL_RE = re.compile(r"^[a-z][a-z0-9]*(_[a-z0-9]+)+$")


def expected_value(label: str, expectations: dict[str, int], stats: dict) -> int | None:
    """Authoritative value for a Summary label, or None if stats cannot speak to it.

    A snake_case label is an exception class. correlate.py only records classes that
    occurred, so a class absent from stats legitimately expects 0 - returning None there
    would let a stale non-zero formula slip through unchecked.
    """
    if label in expectations:
        return expectations[label]
    if CLASS_LABEL_RE.match(label) and isinstance(stats.get("exceptions_by_class"), dict):
        return 0
    return None


def check(ok: bool, name: str, detail: str = "") -> bool:
    RESULTS.append((ok, name, detail))
    return ok


def main() -> int:
    ap = argparse.ArgumentParser(description="Verify a completed review cycle.")
    ap.add_argument("--month", required=True)
    ap.add_argument("--run", type=int, default=1)
    ap.add_argument("--label", default=None)
    args = ap.parse_args()

    paths = month_paths(args.month, args.run)
    label = args.label or args.month
    out, inp = paths["output"], paths["input"]

    stats_path = out / f"correlation-stats-{label}.json"
    if not stats_path.exists():
        print(f"No correlation-stats-{label}.json - nothing to verify.", file=sys.stderr)
        return 2
    stats = json.loads(stats_path.read_text(encoding="utf-8"))

    def read(name):
        p = out / f"{name}-{label}.csv"
        return pd.read_csv(p, keep_default_na=False) if p.exists() and p.stat().st_size > 2 \
            else pd.DataFrame()

    acts, corr, unc, exc = (read("activations"), read("correlated-actions"),
                            read("uncovered-actions"), read("exceptions"))

    # 1 - row count reconciliation
    raw, dedup, dropped = (stats.get("pim_rows_raw", 0), stats.get("pim_rows_deduped", 0),
                           stats.get("pim_exact_duplicates_dropped", 0))
    check(raw == dedup + dropped,
          "PIM row counts reconcile", f"{raw} raw = {dedup} deduped + {dropped} duplicates")
    check(len(acts) == stats.get("activations_successful", -1),
          "Activation count matches stats",
          f"activations csv {len(acts)} vs stats {stats.get('activations_successful')}")
    check(len(acts) <= dedup, "Anchors do not exceed distinct events", f"{len(acts)} <= {dedup}")

    # 1b - audit row reconciliation. Every row leaving the raw file must be accounted for
    # by exactly one of: exact duplicate, unparseable timestamp, chunk fragment collapsed
    # into its parent event, or PIM's own bookkeeping excluded from the action set.
    if stats.get("audit_available"):
        a_after = int(stats.get("audit_rows_after_reassembly", -1))
        if a_after >= 0:
            a_raw = int(stats.get("audit_rows_raw", 0))
            a_dup = int(stats.get("audit_exact_duplicates_dropped", 0))
            a_bad = int(stats.get("audit_rows_unparseable_timestamp", 0))
            a_col = int(stats.get("audit_chunk_rows_collapsed", 0))
            check(a_raw - a_dup - a_bad - a_col == a_after,
                  "Audit row counts reconcile",
                  f"{a_raw} raw - {a_dup} dup - {a_bad} bad ts - {a_col} chunk "
                  f"fragments = {a_after}")
            a_pim = int(stats.get("audit_pim_service_rows_excluded", 0))
            a_elig = int(stats.get("audit_rows_eligible_for_correlation", -1))
            check(a_after - a_pim == a_elig, "Audit eligible count reconciles",
                  f"{a_after} - {a_pim} PIM-service = {a_elig}")
        else:
            check(False, "Audit row counts reconcile",
                  "correlation-stats has no audit_rows_after_reassembly - stats predate "
                  "chunked-event reassembly; re-run correlate.py")

    # 2 - duplicates in outputs
    for name, df in (("activations", acts), ("exceptions", exc),
                     ("correlated-actions", corr), ("uncovered-actions", unc)):
        if df.empty:
            continue
        dupes = int(df.duplicated().sum())
        check(dupes == 0, f"No duplicate rows in {name}", f"{dupes} found")

    # 2b - proof that chunked events were reassembled. A Graph event split across rows
    # gives every fragment the same audit_event_id, and exact-duplicate detection cannot
    # see them because the payload slice and the sub-second Date differ. After reassembly
    # each event is one row, so an id may appear at most once per activation window
    # (across windows is legitimate - that is ambiguous attribution).
    #
    # This replaces an earlier whole-second comparison, which flagged rows that no
    # exported field can distinguish. Those are counted separately and reported rather
    # than failed, because this export omits the field that would tell them apart.
    if not corr.empty and "audit_event_id" in corr.columns:
        withid = corr[corr["audit_event_id"].astype(str).str.strip() != ""]
        n = int(withid.duplicated(subset=["activation_id", "audit_event_id"]).sum())
        check(n == 0, "Chunked audit events reassembled (correlated-actions)",
              f"{n} repeated event id(s) within one activation"
              + ("" if n == 0 else " - fragments were not reassembled"))
    if not unc.empty and "audit_event_id" in unc.columns:
        withid = unc[unc["audit_event_id"].astype(str).str.strip() != ""]
        n = int(withid.duplicated(subset=["audit_event_id"]).sum())
        check(n == 0, "Chunked audit events reassembled (uncovered-actions)",
              f"{n} repeated event id(s)")

    if not acts.empty:
        d = int(acts["activation_id"].duplicated().sum())
        check(d == 0, "activation_id is unique", f"{d} duplicates")
    if not exc.empty:
        d = int(exc["exception_id"].duplicated().sum())
        check(d == 0, "exception_id is unique", f"{d} duplicates")

    # 3 - exception totals vs Summary formulas
    xlsx = out / f"entra-pim-correlation-{label}.xlsx"
    if xlsx.exists():
        try:
            from openpyxl import load_workbook
            wb = load_workbook(xlsx, data_only=True)
            summary = {}
            ws = wb["Summary"]
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=2):
                a, b = row[0].value, row[1].value
                if a is not None and b not in (None, ""):
                    summary[str(a).strip()] = b
            # The Summary sheet holds live formulas by design (CLAUDE.md), and openpyxl
            # writes formulas without computing them - so data_only=True yields None for
            # every formula cell until Excel or LibreOffice has opened the file. These
            # checks must therefore NOT depend on a cached value being present; when they
            # did, they silently passed without comparing anything.
            #
            # correlation-stats-<label>.json is the authority for every figure the reports
            # use, so each Summary formula is recomputed from the CSVs and compared against
            # stats. That works with or without cached values, and additionally catches a
            # formula pointing at the wrong range or criterion.
            check(int(stats.get("exception_rows", -1)) == len(exc),
                  "Exception total matches Exceptions sheet",
                  f"stats {stats.get('exception_rows')} vs csv {len(exc)}")
            if not exc.empty:
                stats_sev = stats.get("exceptions_by_severity") or {}
                csv_sev = {s: int((exc["severity"] == s).sum())
                           for s in exc["severity"].unique()}
                mismatched = {s: (stats_sev.get(s), c) for s, c in csv_sev.items()
                              if int(stats_sev.get(s, -1)) != c}
                check(not mismatched, "Severity counts match Exceptions sheet",
                      f"stats vs csv: {mismatched}" if mismatched else
                      ", ".join(f"{s}={c}" for s, c in sorted(csv_sev.items())))
                by_sev = sum(csv_sev.values())
                check(by_sev == len(exc), "Severity buckets sum to total",
                      f"{by_sev} vs {len(exc)}")

            # Re-evaluate every Summary formula against the source CSVs, then compare to
            # stats (always available) and to the cached value (only after a recalculation).
            frames = {"Activations": acts, "Exceptions": exc,
                      "Correlated Actions": corr, "Uncovered Actions": unc}
            expected_for = summary_expectations(stats)
            wbf = load_workbook(xlsx, data_only=False)["Summary"]
            evaluated = wrong = compared = cached_hits = unmapped = 0
            for row in wbf.iter_rows(min_row=1, max_row=wbf.max_row, max_col=2):
                lbl, f = row[0].value, row[1].value
                if not isinstance(f, str) or not f.startswith("="):
                    continue
                recomputed = eval_count_formula(f, frames)
                if recomputed is None:
                    continue
                evaluated += 1
                name = str(lbl).strip()

                authority = expected_value(name, expected_for, stats)
                if authority is not None:
                    compared += 1
                    if int(authority) != recomputed:
                        wrong += 1
                        check(False, f"Summary formula disagrees with stats: {name}",
                              f"stats {authority} vs formula recomputed from CSV {recomputed}")
                else:
                    unmapped += 1

                cached = summary.get(name)
                if isinstance(cached, (int, float)):
                    cached_hits += 1
                    if int(cached) != recomputed:
                        wrong += 1
                        check(False, f"Cached formula value wrong: {name}",
                              f"workbook {cached} vs recomputed {recomputed}")

            detail = (f"{compared} of {evaluated} formulas checked against "
                      f"correlation-stats-{label}.json")
            if cached_hits:
                detail += f"; {cached_hits} also matched cached values"
            if unmapped:
                detail += f"; {unmapped} had no stats counterpart"
            # Passing requires that something was actually compared. A run where every
            # formula was unverifiable is not a pass.
            check(wrong == 0 and compared > 0, "Summary formulas agree with stats", detail)
            if unmapped:
                print(f"  NOTE  {unmapped} Summary formula(s) have no counterpart in "
                      f"correlation-stats-{label}.json and were recomputed but not "
                      f"cross-checked.")
            if not cached_hits and evaluated:
                print("  NOTE  Summary formulas carry no cached values (openpyxl writes "
                      "formulas without computing them). Excel computes them on open; the "
                      "checks above used correlation-stats instead and did not depend on "
                      "cached values.")
            sheets = set(wb.sheetnames)
            expected = {"Summary", "Activations", "Correlated Actions",
                        "Unmatched Activations", "Uncovered Actions", "Exceptions",
                        "Decisions", "Evidence"}
            check(expected <= sheets, "All expected sheets present",
                  f"missing: {sorted(expected - sheets)}")
        except Exception as exc_err:
            check(False, "Workbook readable", str(exc_err)[:200])
    else:
        check(False, "Workbook exists", f"{xlsx.name} not found")

    # 4 - manifest coverage
    manifest = read_manifest(inp)
    listed = {e.get("file") for e in manifest.get("entries", [])}
    csvs = {p.name for p in inp.glob("*.csv")}
    missing = csvs - listed
    check(not missing, "Every input CSV is in export-manifest.json",
          f"unlisted: {sorted(missing)}" if missing else "")

    # 5 - referential integrity
    if not corr.empty and not acts.empty:
        orphans = set(corr["activation_id"]) - set(acts["activation_id"])
        check(not orphans, "Correlated rows reference real activations",
              f"orphans: {sorted(orphans)[:5]}")
    if not exc.empty and not acts.empty:
        used = {a for a in exc["activation_id"].astype(str) if a and a != "nan"}
        orphans = used - set(acts["activation_id"])
        check(not orphans, "Exception rows reference real activations",
              f"orphans: {sorted(orphans)[:5]}")

    # 6 - every correlated action inside its window
    if not corr.empty:
        c = corr.copy()
        for col in ("activation_utc", "window_end_utc", "audit_Date"):
            c[col] = pd.to_datetime(c[col], errors="coerce", utc=True, format="mixed")
        bad = int(((c["audit_Date"] < c["activation_utc"]) |
                   (c["audit_Date"] >= c["window_end_utc"])).sum())
        check(bad == 0, "Every correlated action falls inside its window",
              f"{bad} outside")
        neg = int((corr["minutes_after_activation"] < 0).sum())
        check(neg == 0, "No negative minutes-after-activation", f"{neg} negative")

    # 7 - integrity of the no-audit case
    if not stats.get("audit_available"):
        if not acts.empty:
            statuses = set(acts["correlation_status"].unique())
            check(statuses == {"unknown - no audit data"},
                  "No-audit cycle labels every activation 'unknown'",
                  f"found: {sorted(statuses)}")
        if not exc.empty:
            bad = int((exc["exception_class"] == "activation_no_actions").sum())
            check(bad == 0, "No 'unused activation' findings without audit data",
                  f"{bad} such rows - would be unsupported by evidence")

    # report
    width = max(len(n) for _, n, _ in RESULTS) + 2
    print(f"\nVerification - {args.month} (label {label})\n{'-' * (width + 12)}")
    failures = 0
    for ok, name, detail in RESULTS:
        failures += not ok
        print(f"  {'PASS' if ok else 'FAIL'}  {name.ljust(width)}{detail}")
    print(f"{'-' * (width + 12)}\n  {len(RESULTS) - failures}/{len(RESULTS)} passed"
          f"{'' if not failures else f' - {failures} FAILURE(S)'}")

    if not stats.get("audit_available"):
        print("\n  NOTE: no directory audit export was present, so the correlation checks "
              "that depend on it were not exercised.")

    # Not a failure - these may be real repeated actions - but the ambiguity has to be
    # visible, because it bounds how precisely any action count can be read.
    ci = int(stats.get("audit_content_identical_rows", 0) or 0)
    if ci:
        print(f"\n  NOTE: {ci} audit row(s) are identical to another row in every exported "
              f"field except the sub-second timestamp. They are kept, not deduplicated: "
              f"the field that would distinguish them (modifiedProperties) is not in this "
              f"export, so action volume for those events is an upper bound.")

    # Not a failure - a new Graph event type should not block a cycle - but it must be
    # visible in the gate output, because unmapped rows generate no findings at all.
    unmapped = stats.get("pim_unmapped_actions") or {}
    if unmapped:
        print(f"\n  NOTE: {stats.get('pim_unmapped_action_rows', 0)} PIM row(s) carry an "
              f"action correlate.py does not map and so produce no findings:")
        for name, n in sorted(unmapped.items(), key=lambda kv: -kv[1]):
            print(f"          {name} = {n}")
    return 1 if failures else 0


if __name__ == "__main__":
    raise SystemExit(main())
